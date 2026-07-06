from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

NAVY = "1B2A4A"
GOLD = "C9A84C"
LIGHT_GOLD = "F5EFE0"
LIGHT_GRAY = "F0F0F0"
WHITE = "FFFFFF"
RED = "CC0000"
GREEN = "228B22"
INPUT_BG = "FFFFF0"

navy_fill = PatternFill("solid", fgColor=NAVY)
gold_fill = PatternFill("solid", fgColor=LIGHT_GOLD)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
input_fill = PatternFill("solid", fgColor=INPUT_BG)
white_fill = PatternFill("solid", fgColor=WHITE)

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
title_font = Font(name="Arial", bold=True, color=WHITE, size=14)
section_font = Font(name="Arial", bold=True, color=NAVY, size=11)
gold_font = Font(name="Arial", bold=True, color=GOLD, size=11)
normal_font = Font(name="Arial", color="333333", size=10)
bold_font = Font(name="Arial", bold=True, color="333333", size=10)
result_font = Font(name="Arial", bold=True, color=NAVY, size=11)
red_font = Font(name="Arial", bold=True, color=RED, size=11)
green_font = Font(name="Arial", bold=True, color=GREEN, size=11)
pct_font = Font(name="Arial", bold=True, color=NAVY, size=12)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def style_range(ws, row, cols, font, fill, border=thin_border, align=None):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = align or Alignment(horizontal="center", vertical="center")

def set_cell(ws, row, col, value, font=normal_font, fill=white_fill, fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.border = thin_border
    cell.alignment = align or Alignment(horizontal="right" if isinstance(value, (int, float)) else "left", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell

# ==================== HOJA 1: Unit Economics ====================
ws1 = wb.active
ws1.title = "Unit Economics"
ws1.sheet_properties.tabColor = NAVY

for c in range(1, 6):
    ws1.column_dimensions[get_column_letter(c)].width = [2, 32, 20, 20, 20][c-1]

ws1.merge_cells("B1:E1")
ws1.merge_cells("B2:E2")
set_cell(ws1, 1, 2, "BRICKTOKEN - Unit Economics por Propiedad", title_font, navy_fill, align=Alignment(horizontal="center", vertical="center"))
set_cell(ws1, 2, 2, "Modelo financiero por tipo de propiedad", Font(name="Arial", color=GOLD, size=10, italic=True), navy_fill, align=Alignment(horizontal="center", vertical="center"))
for c in [3,4,5]:
    ws1.cell(row=1, column=c).fill = navy_fill
    ws1.cell(row=2, column=c).fill = navy_fill

r = 4
headers = ["", "Concepto", "Monoambiente Cordon", "Local Comercial", "Apto 2 Dorm Pocitos"]
for c, h in enumerate(headers, 1):
    set_cell(ws1, r, c, h, header_font, navy_fill, align=Alignment(horizontal="center", vertical="center", wrap_text=True))

# Data sections
data = [
    ("section", "DATOS DE LA PROPIEDAD"),
    ("data", "Valor de compra (USD)", 120000, 350000, 360000, "$#,##0"),
    ("data", "Alquiler mensual (USD)", 750, 3500, 1490, "$#,##0"),
    ("data", "Precio del token (USD)", 100, 100, 100, "$#,##0"),
    ("formula", "Total tokens", "=C6/C9", "=D6/D9", "=E6/E9", "#,##0"),
    ("section", "COSTOS ANUALES"),
    ("data", "Contribucion + Primaria", 800, 2500, 2500, "$#,##0"),
    ("data", "Mantenimiento (3%)", None, None, None, "$#,##0", "=C7*12*0.03", "=D7*12*0.03", "=E7*12*0.03"),
    ("data", "Vacancia (3-5%)", None, None, None, "$#,##0", "=C7*12*0.03", "=D7*12*0.05", "=E7*12*0.05"),
    ("data", "Seguro", 200, 400, 400, "$#,##0"),
    ("formula", "Total costos", "=SUM(C13:C16)", "=SUM(D13:D16)", "=SUM(E13:E16)", "$#,##0"),
    ("section", "INGRESOS"),
    ("formula", "Ingreso bruto anual", "=C7*12", "=D7*12", "=E7*12", "$#,##0"),
    ("formula", "Ingreso neto", "=C20-C17", "=D20-D17", "=E20-E17", "$#,##0"),
    ("section", "FEE BRICKTOKEN"),
    ("formula", "Management fee (8%)", "=C21*0.08", "=D21*0.08", "=E21*0.08", "$#,##0"),
    ("section", "PARA EL INVERSOR"),
    ("formula", "Distribucion anual", "=C21-C23", "=D21-D23", "=E21-E23", "$#,##0"),
    ("formula", "Yield neto (%)", "=C25/C6", "=D25/D6", "=E25/E6", "0.0%"),
    ("data", "Apreciacion estimada", "3-5%", "2-4%", "3-5%", None),
    ("data", "Retorno total estimado", "8.7-10.7%", "11.4-13.4%", "6.5-8.5%", None),
    ("section", "POR TOKEN"),
    ("formula", "Distribucion anual/token", "=C25/C10", "=D25/D10", "=E25/E10", "$#,##0.00"),
    ("formula", "Distribucion trimestral/token", "=C30/4", "=D30/4", "=E30/4", "$#,##0.00"),
]

r = 5
for item in data:
    if item[0] == "section":
        r += 1
        set_cell(ws1, r, 2, item[1], section_font, gold_fill, align=Alignment(horizontal="left"))
        for c in [3, 4, 5]:
            set_cell(ws1, r, c, "", section_font, gold_fill)
    elif item[0] == "data":
        r += 1
        label, v1, v2, v3, fmt = item[1], item[2], item[3], item[4], item[5]
        is_input = isinstance(v1, (int, float))
        f = bold_font if "Total" in label else normal_font
        bg = gray_fill if r % 2 == 0 else white_fill
        set_cell(ws1, r, 2, label, f, bg, align=Alignment(horizontal="left"))
        if len(item) > 6:
            set_cell(ws1, r, 3, item[6], f, bg, fmt)
            set_cell(ws1, r, 4, item[7], f, bg, fmt)
            set_cell(ws1, r, 5, item[8], f, bg, fmt)
        else:
            set_cell(ws1, r, 3, v1, f, bg, fmt)
            set_cell(ws1, r, 4, v2, f, bg, fmt)
            set_cell(ws1, r, 5, v3, f, bg, fmt)
    elif item[0] == "formula":
        r += 1
        label, f1, f2, f3, fmt = item[1], item[2], item[3], item[4], item[5]
        is_result = "Distribucion anual" == label or "Yield" in label
        f = result_font if is_result else bold_font
        bg = gold_fill if is_result else (gray_fill if r % 2 == 0 else white_fill)
        set_cell(ws1, r, 2, label, f, bg, align=Alignment(horizontal="left"))
        set_cell(ws1, r, 3, f1, f, bg, fmt)
        set_cell(ws1, r, 4, f2, f, bg, fmt)
        set_cell(ws1, r, 5, f3, f, bg, fmt)

# ==================== HOJA 2: Revenue BrickToken ====================
ws2 = wb.create_sheet("Revenue BrickToken")
ws2.sheet_properties.tabColor = GOLD

for c in range(1, 6):
    ws2.column_dimensions[get_column_letter(c)].width = [2, 36, 18, 18, 18][c-1]

ws2.merge_cells("B1:E1")
ws2.merge_cells("B2:E2")
set_cell(ws2, 1, 2, "BRICKTOKEN - Proyecciones de Revenue", title_font, navy_fill, align=Alignment(horizontal="center"))
set_cell(ws2, 2, 2, "Modelo a 3 anos", Font(name="Arial", color=GOLD, size=10, italic=True), navy_fill, align=Alignment(horizontal="center"))
for c in [3,4,5]:
    ws2.cell(row=1, column=c).fill = navy_fill
    ws2.cell(row=2, column=c).fill = navy_fill

r = 4
for c, h in enumerate(["", "Concepto", "Ano 1", "Ano 2", "Ano 3"], 1):
    set_cell(ws2, r, c, h, header_font, navy_fill, align=Alignment(horizontal="center"))

rev_data = [
    ("section", "ESCALA"),
    ("input", "Propiedades activas", 4, 15, 40),
    ("input", "Valor promedio propiedad (USD)", 200000, 200000, 200000),
    ("formula", "Capital tokenizado total (AUM)", "=C6*C7", "=D6*D7", "=E6*E7", "$#,##0"),
    ("input", "% tokens vendidos", 0.60, 0.70, 0.80),
    ("formula", "Capital vendido a inversores", "=C8*C9", "=D8*D9", "=E8*E9", "$#,##0"),
    ("input", "Inversores activos", 50, 300, 1200),
    ("section", "REVENUE"),
    ("formula", "Fee compra (2.5%)", "=C10*0.025", "=(D10-C10)*0.025", "=(E10-D10)*0.025", "$#,##0"),
    ("formula", "Management fee anual (8% neto)", "=C8*0.05*0.08", "=D8*0.05*0.08", "=E8*0.05*0.08", "$#,##0"),
    ("formula", "Fee marketplace (1.5%)", 0, "=D10*0.03*0.015", "=E10*0.05*0.015", "$#,##0"),
    ("formula", "Renta tokens propios", "=(C8-C10)*0.05", "=(D8-D10)*0.05", "=(E8-E10)*0.05", "$#,##0"),
    ("formula", "Revenue total", "=SUM(C13:C16)", "=SUM(D13:D16)", "=SUM(E13:E16)", "$#,##0"),
    ("section", "COSTOS OPERATIVOS"),
    ("input", "Legal/fideicomiso ($3K x prop)", None, None, None, "=C6*3000", "=D6*3000", "=E6*3000"),
    ("input", "Desarrollo + hosting", 24000, 36000, 60000),
    ("input", "Compliance/contabilidad", 12000, 18000, 24000),
    ("input", "Marketing", 6000, 24000, 48000),
    ("input", "Equipo (salarios)", 0, 36000, 120000),
    ("formula", "Tasaciones ($400 x prop x 2/ano)", "=C6*400*2", "=D6*400*2", "=E6*400*2", "$#,##0"),
    ("formula", "Total costos", "=SUM(C19:C24)", "=SUM(D19:D24)", "=SUM(E19:E24)", "$#,##0"),
    ("section", "RESULTADO"),
    ("formula", "EBITDA", "=C17-C25", "=D17-D25", "=E17-E25", "$#,##0;($#,##0);-"),
    ("formula", "Margen", "=IF(C17=0,0,C27/C17)", "=IF(D17=0,0,D27/D17)", "=IF(E17=0,0,E27/E17)", "0.0%"),
]

r = 5
for item in rev_data:
    if item[0] == "section":
        r += 1
        set_cell(ws2, r, 2, item[1], section_font, gold_fill, align=Alignment(horizontal="left"))
        for c in [3,4,5]:
            set_cell(ws2, r, c, "", section_font, gold_fill)
    elif item[0] == "input":
        r += 1
        label = item[1]
        bg = gray_fill if r % 2 == 0 else white_fill
        set_cell(ws2, r, 2, label, normal_font, bg, align=Alignment(horizontal="left"))
        fmt_str = "$#,##0" if "USD" in label or "hosting" in label.lower() or "compliance" in label.lower() or "marketing" in label.lower() or "equipo" in label.lower() or "legal" in label.lower() else ("#,##0" if isinstance(item[2], int) and item[2] > 1 else "0.0%")
        if len(item) > 5:
            set_cell(ws2, r, 3, item[5], normal_font, bg, "$#,##0")
            set_cell(ws2, r, 4, item[6], normal_font, bg, "$#,##0")
            set_cell(ws2, r, 5, item[7], normal_font, bg, "$#,##0")
        else:
            set_cell(ws2, r, 3, item[2], normal_font, bg, fmt_str)
            set_cell(ws2, r, 4, item[3], normal_font, bg, fmt_str)
            set_cell(ws2, r, 5, item[4], normal_font, bg, fmt_str)
    elif item[0] == "formula":
        r += 1
        label, f1, f2, f3, fmt = item[1], item[2], item[3], item[4], item[5]
        is_total = "total" in label.lower() or "EBITDA" in label or "Margen" in label
        f = result_font if is_total else bold_font
        bg = gold_fill if is_total else (gray_fill if r % 2 == 0 else white_fill)
        if "EBITDA" in label:
            f = red_font
        if "Margen" in label:
            f = bold_font
        set_cell(ws2, r, 2, label, f, bg, align=Alignment(horizontal="left"))
        set_cell(ws2, r, 3, f1, f, bg, fmt)
        set_cell(ws2, r, 4, f2, f, bg, fmt)
        set_cell(ws2, r, 5, f3, f, bg, fmt)

# ==================== HOJA 3: Simulador Inversor ====================
ws3 = wb.create_sheet("Simulador Inversor")
ws3.sheet_properties.tabColor = "228B22"

for c in range(1, 9):
    ws3.column_dimensions[get_column_letter(c)].width = [2, 30, 18, 18, 18, 18, 18, 18][c-1]

ws3.merge_cells("B1:G1")
ws3.merge_cells("B2:G2")
set_cell(ws3, 1, 2, "SIMULADOR DE INVERSION - BRICKTOKEN", title_font, navy_fill, align=Alignment(horizontal="center"))
set_cell(ws3, 2, 2, "Ingresa tus datos y calcula tu retorno estimado", Font(name="Arial", color=GOLD, size=10, italic=True), navy_fill, align=Alignment(horizontal="center"))
for c in range(3, 8):
    ws3.cell(row=1, column=c).fill = navy_fill
    ws3.cell(row=2, column=c).fill = navy_fill

r = 4
set_cell(ws3, r, 2, "INPUTS (modifica estos valores)", section_font, gold_fill, align=Alignment(horizontal="left"))
set_cell(ws3, r, 3, "", section_font, gold_fill)

inputs = [
    ("Monto a invertir (USD)", 1000, "$#,##0"),
    ("Yield neto anual (%)", 0.057, "0.0%"),
    ("Apreciacion anual (%)", 0.04, "0.0%"),
    ("Fee de compra (%)", 0.025, "0.0%"),
    ("Horizonte (anos)", 5, "#,##0"),
]

for i, (label, val, fmt) in enumerate(inputs):
    r += 1
    set_cell(ws3, r, 2, label, bold_font, input_fill, align=Alignment(horizontal="left"))
    c = set_cell(ws3, r, 3, val, Font(name="Arial", bold=True, color="0000FF", size=12), input_fill, fmt)

r += 2
set_cell(ws3, r, 2, "Inversion efectiva (post fee)", result_font, gold_fill, align=Alignment(horizontal="left"))
set_cell(ws3, r, 3, "=C5*(1-C8)", result_font, gold_fill, "$#,##0.00")

r += 2
headers = ["Ano", "Valor Tokens", "Renta del Ano", "Renta Acumulada", "Valor Total", "Retorno (%)"]
for c, h in enumerate(headers, 2):
    set_cell(ws3, r, c, h, header_font, navy_fill, align=Alignment(horizontal="center"))

inv_cell = "$C$11"  # inversion efectiva
yield_cell = "$C$6"
apre_cell = "$C$7"

for yr in range(1, 11):
    r += 1
    bg = gray_fill if yr % 2 == 0 else white_fill
    set_cell(ws3, r, 2, yr, bold_font, bg, "#,##0", align=Alignment(horizontal="center"))
    # Valor tokens = inversion efectiva * (1 + apreciacion)^ano
    set_cell(ws3, r, 3, f"={inv_cell}*(1+{apre_cell})^B{r}", normal_font, bg, "$#,##0.00")
    # Renta del ano = inversion efectiva * yield (sobre valor original)
    set_cell(ws3, r, 4, f"={inv_cell}*{yield_cell}", normal_font, bg, "$#,##0.00")
    # Renta acumulada
    if yr == 1:
        set_cell(ws3, r, 5, f"=D{r}", normal_font, bg, "$#,##0.00")
    else:
        set_cell(ws3, r, 5, f"=E{r-1}+D{r}", normal_font, bg, "$#,##0.00")
    # Valor total = valor tokens + renta acumulada
    set_cell(ws3, r, 6, f"=C{r}+E{r}", bold_font, bg, "$#,##0.00")
    # Retorno %
    set_cell(ws3, r, 7, f"=(F{r}-$C$5)/$C$5", bold_font, bg, "0.0%")

# ==================== HOJA 4: Fees & Comparativa ====================
ws4 = wb.create_sheet("Fees & Comparativa")
ws4.sheet_properties.tabColor = GOLD

for c in range(1, 7):
    ws4.column_dimensions[get_column_letter(c)].width = [2, 24, 20, 20, 20, 20][c-1]

ws4.merge_cells("B1:F1")
ws4.merge_cells("B2:F2")
set_cell(ws4, 1, 2, "COMPARATIVA DE INVERSIONES", title_font, navy_fill, align=Alignment(horizontal="center"))
set_cell(ws4, 2, 2, "BrickToken vs. alternativas tradicionales", Font(name="Arial", color=GOLD, size=10, italic=True), navy_fill, align=Alignment(horizontal="center"))
for c in range(3, 7):
    ws4.cell(row=1, column=c).fill = navy_fill
    ws4.cell(row=2, column=c).fill = navy_fill

r = 4
comp_headers = ["Concepto", "BrickToken", "Compra directa", "Plazo fijo USD", "S&P 500"]
for c, h in enumerate(comp_headers, 2):
    fill = navy_fill
    if h == "BrickToken":
        fill = PatternFill("solid", fgColor=GOLD)
    set_cell(ws4, r, c, h, Font(name="Arial", bold=True, color=WHITE if h != "BrickToken" else NAVY, size=11), fill, align=Alignment(horizontal="center"))

comp_data = [
    ("Inversion minima", "$100", "$120,000+", "$1,000", "$500"),
    ("Liquidez", "Media (marketplace)", "Muy baja (meses)", "Alta (30-365 dias)", "Alta (inmediata)"),
    ("Gestion necesaria", "Ninguna", "Total", "Ninguna", "Ninguna"),
    ("Yield estimado", "5-9%", "4-6%", "~3%", "~10%"),
    ("Respaldo tangible", "Inmueble real", "Inmueble real", "Banco", "Acciones"),
    ("Diversificacion", "Multiples props", "1 propiedad", "N/A", "Alta"),
    ("Transparencia", "Total (docs, tasaciones)", "Variable", "Baja", "Alta"),
    ("Fees", "2.5% entrada + 8% mgmt", "Escribania ~3% + imp.", "0%", "0.03-0.5%"),
    ("Accesibilidad", "Online, inmediato", "Presencial, meses", "Presencial/online", "Online"),
    ("Proteccion inflacion", "Si (inmueble)", "Si (inmueble)", "No", "Parcial"),
]

for i, (concepto, *vals) in enumerate(comp_data):
    r += 1
    bg = gray_fill if i % 2 == 0 else white_fill
    set_cell(ws4, r, 2, concepto, bold_font, bg, align=Alignment(horizontal="left"))
    for c, v in enumerate(vals, 3):
        f = bold_font if c == 3 else normal_font
        set_cell(ws4, r, c, v, f, gold_fill if c == 3 else bg, align=Alignment(horizontal="center"))

# Save
output_path = "/Users/enriquezerbino/Desktop/bricktoken/docs/BrickToken_Modelo_Financiero.xlsx"
wb.save(output_path)
print(f"Excel saved to {output_path}")
